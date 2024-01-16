from twilio.rest import Client
import openpyxl as opxl

# Input the twilio provided data
my_phone_number = "+123456789"
auth_token = "######################"
account_SID = "######################"

client = Client(account_SID, auth_token)

# Connect the script to your excel database
load_workbook = opxl.load_workbook("PATH")
sheet = load_workbook.active

row_qantity = sheet.max_row


def sms_handler():
    starting_row = 2
    for i in range(row_qantity - 1):

        # Adjust the cell names to fit your needs
        # Mine is set to work with data I need
        def name_handler():
            name_value = sheet[f"{'A'}{starting_row}"].value
            return name_value

        def number_handler():
            number_value = sheet[f"{'B'}{starting_row}"].value
            return number_value

        def code_handler():
            code_value = sheet[f"{'C'}{starting_row}"].value
            return code_value

        # This gets work done :)
        # Again, set the message how you want it to be
        message = client.messages.create(
            body=f"Greetings {name_handler()}! Your code for this month is {code_handler()}. Use it well!",
            from_=my_phone_number,
            to=f"+{number_handler()}",
        )

        # This gives you feedback in console to check for errors or completion
        print(message.body)
        print(f"Message sent to number: {number_handler()}. SID: {message.sid}")

        # Increment the row count
        starting_row += 1


sms_handler()
